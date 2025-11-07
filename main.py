from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import re

from utils import is_valid_file_path, has_duplicates, get_uppercase_list

#|--------------------------------------------------------------|
#| INTERNAL FUNCTIONS:
#|--------------------------------------------------------------|
def _get_cel_coordinate(ws, value) -> str:
    '''
    Pegar a posição de uma determinada celula na tabela.
    '''
    for row in ws.iter_rows():
        for cel in row:
            if cel.value == value:
                return cel.coordinate

    raise FileNotFoundError('Valor da celula não encontrado na tabela.')

#|--------------------------------------------------------------|
def _is_col_none(ws, col_letter: str) -> bool:
    '''
    Verifica se a coluna está vazia.
    '''
    for row in range(1, ws.max_row + 1):
        if ws[f'{col_letter}{row}'].value:
            return False
    
    return True

#|--------------------------------------------------------------|
def _is_row_none(ws, row: int) -> bool:
    '''
    Verifica se a linha está vazia.
    '''
    for col_index in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_index)

        if ws[f'{col_letter}{row}'].value:
            return False
    
    return True

#|--------------------------------------------------------------|
def _get_col_values(
        ws,
        header_value,
        end_row: int = None
) -> list:
    '''
    Pega todos os valores de uma coluna a partir de um cabeçalho.
    '''
    cel_header = _get_cel_coordinate(ws, header_value)
    col_letter = re.sub(r'\d', '', cel_header)
    col_number = int(re.sub(r'[^0-9]', '', cel_header))
    col_values = []

    if not end_row:
        end_row = ws.max_row

    for row in range(col_number + 1, end_row + 1):
        cel_value = ws[f'{col_letter}{row}'].value

        if _is_row_none(ws, row):
            continue

        col_values.append(cel_value)

    return col_values

#|--------------------------------------------------------------|
def _get_row_values(
        ws,
        header_value,
        end_col: int = None
) -> list:
    '''
    Pega todos os valores de uma linha a partir de um cabeçalho.
    '''
    cel_header = ws[_get_cel_coordinate(ws, header_value)]
    col_values = []

    if not end_col:
        end_col = ws.max_column

    for col_index in range(cel_header.column + 1, end_col + 1): 
        col_letter = get_column_letter(col_index)
        col_value = ws[f'{col_letter}{cel_header.row}'].value

        if _is_col_none(ws, col_letter):
            continue
        
        col_values.append(col_value)

    return col_values

#|--------------------------------------------------------------|
#| MAIN FUNCTIONS:
#|--------------------------------------------------------------|
# Tabela organizada em "COLUNAS":
def spreadsheet_cols_to_dict(
        file_path: str,
        ws_name: str,
        headers: list,
        end_row: int = None
) -> dict:
    '''
    Recebe um arquivo do tipo xlsx, xls, Excel ou arquivos CSV e converte para um dicionário. Identifica automaticamente os dados de acordo com a posição dos cabeçalhos.
    '''
    allowed_extensions = ['xlsx', 'xls', 'csv']
    new_dict = {}

    # Validações:
    if not is_valid_file_path(file_path, allowed_extensions):
        raise FileNotFoundError('Caminho do arquivo ou extensão invalida.')
    
    if not headers:
        raise ValueError('É necessario informar ao menos um dos cabeçalhos (headers).')
    
    if has_duplicates(headers):
        raise ValueError('Não pode haver mais de um cabeçalho (header) com o mesmo nome.')
    
    # Buscar dados:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[ws_name]

    for header in headers:
            new_dict[str(header).upper()] = _get_col_values(ws, header, end_row)

    # Prerações:
    for header in new_dict:
        new_dict[header] = get_uppercase_list(new_dict[header])

    return new_dict

#|--------------------------------------------------------------|
# Tabela organizada em "LINHAS":
def spreadsheet_rows_to_dict(
        file_path: str,
        ws_name: str,
        headers: list,
        end_col: int = None
) -> dict:
    '''
    Recebe um arquivo do tipo xlsx, xls, Excel ou arquivos CSV e converte para um dicionário. Identifica automaticamente os dados de acordo com a posição dos cabeçalhos.
    '''
    allowed_extensions = ['xlsx', 'xls', 'csv']
    new_dict = {}

    # Validações:
    if not is_valid_file_path(file_path, allowed_extensions):
        raise FileNotFoundError('Caminho do arquivo ou extensão invalida.')
    
    if not headers:
        raise ValueError('É necessario informar ao menos um dos cabeçalhos (headers).')
    
    if has_duplicates(headers):
        raise ValueError('Não pode haver mais de um cabeçalho (header) com o mesmo nome.')
    
    # Buscar dados:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[ws_name]

    for header in headers:
            new_dict[str(header).upper()] = _get_row_values(ws, header, end_col)

    # Prerações:
    for header in new_dict:
        new_dict[header] = get_uppercase_list(new_dict[header])

    return new_dict

#|--------------------------------------------------------------|
#| TESTS:
#|--------------------------------------------------------------|
if __name__ == '__main__':
    # Configurações:
    file_path = 'TESTE.xlsx'
    sheet_name = 'teste'
    horizontal_headers = ['Nome', 'Valor']
    vertical_headers = ['Total']

    # Execução:
    dict_data = spreadsheet_cols_to_dict(file_path, sheet_name, horizontal_headers, 7)
    dict_totals = spreadsheet_rows_to_dict(file_path, sheet_name, vertical_headers)

    print(dict_data, '\n')
